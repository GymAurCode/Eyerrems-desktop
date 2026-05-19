"""Bulk import API — templates, validate, execute, history."""
import csv
import io
import json
import zipfile
from datetime import datetime
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_roles
from app.core.database import get_db
from app.models.auth import User
from app.models.import_batch import ImportBatch, ImportRowLog
from app.schemas.import_schema import (
    ImportBatchOut,
    ImportExecuteRequest,
    ImportExecuteResponse,
    ImportModuleOut,
    ImportRowPreview,
    ImportValidateResponse,
    MasterValidateResponse,
    MasterExecuteRequest,
    MasterExecuteResponse,
)
from app.services.bulk_import.engine import create_batch, execute_import, validate_rows
from app.services.bulk_import.parser import parse_upload
from app.services.bulk_import.registry import import_registry
from app.services.bulk_import.template_service import build_template
from app.services.bulk_import.types import ImportContext

router = APIRouter()

# In-memory validation cache: batch_id -> (rows, module_key, duplicate_mode)
_validation_cache: dict[int, tuple[list[dict[str, str]], str, str]] = {}


def _company_id(user: User) -> int | None:
    return None if user.is_super_admin else user.company_id


def _user_can_import(user: User, permission: str | None) -> bool:
    if user.is_super_admin:
        return True
    if _is_admin(user):
        return True
    if not permission:
        return True
    perms = set()
    for role in user.roles or []:
        for p in role.permissions or []:
            perms.add(p.name)
    if user.role:
        for p in user.role.permissions or []:
            perms.add(p.name)
    for p in user.direct_permissions or []:
        perms.add(p.name)
    return permission in perms


def _is_admin(user: User) -> bool:
    for role in user.roles or []:
        if role.name.lower() == "admin":
            return True
    if user.role and user.role.name.lower() == "admin":
        return True
    return False


def _get_handler(module_key: str):
    handler = import_registry.get(module_key)
    if not handler:
        raise HTTPException(404, f"Unknown import module: {module_key}")
    return handler


@router.get("/modules", response_model=list[ImportModuleOut])
def list_modules(
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager", "Staff", "Accountant")),
):
    out = []
    for h in import_registry.list_modules():
        if _user_can_import(user, h.permission):
            out.append(ImportModuleOut(**import_registry.module_dict(h)))
    return out


@router.get("/templates/combined")
def download_combined_template(
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager", "Staff", "Accountant")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet configs
    sheets_config = {
        "Leads": [
            ("Name", "John Doe (Required)", "blue"),
            ("Phone", "+1234567890 (Required)", "blue"),
            ("Email", "john.doe@example.com (Optional)", "blue"),
            ("Source", "Website (Optional)", "blue"),
            ("Status", "new (Optional)", "blue"),
            ("Assigned To", "Staff member name (Optional)", "blue"),
        ],
        "Properties": [
            ("Property Title", "Sunset Villa (Required)", "emerald"),
            ("Type", "House (Required)", "emerald"),
            ("Price", "15000000 (Required)", "emerald"),
            ("Location", "DHA Phase 6 (Required)", "emerald"),
            ("Status", "available (Optional)", "emerald"),
            ("Owner", "John Smith (Optional)", "emerald"),
        ],
        "Employees": [
            ("Full Name", "Ali Hassan (Required)", "purple"),
            ("CNIC", "12345-6789012-3 (Required)", "purple"),
            ("Phone", "+923001112233 (Required)", "purple"),
            ("Department", "Sales (Optional)", "purple"),
            ("Position", "Sales Executive (Optional)", "purple"),
            ("Salary", "75000 (Optional)", "purple"),
        ],
        "Town Units": [
            ("Town Name", "Green Valley (Required)", "emerald"),
            ("Block Name", "Block A (Required)", "emerald"),
            ("Unit Number", "P-101 (Required)", "emerald"),
            ("Unit Type", "plot (Required - plot, house, apartment, flat, shop, office, warehouse, farmhouse, etc.)", "emerald"),
            ("Category", "residential (Required - residential, commercial, mixed_use, industrial)", "emerald"),
            ("Size", "5 Marla (Optional)", "emerald"),
            ("Total Price", "5000000 (Optional)", "emerald"),
            ("Status", "available (Optional - available, booked, sold, rented, under_construction, inactive)", "emerald"),
        ],
        "Shops": [
            ("Town Name", "Green Valley (Required)", "emerald"),
            ("Block Name", "Block A (Required)", "emerald"),
            ("Unit Number", "S-101 (Required)", "emerald"),
            ("Size", "5 Marla (Optional)", "emerald"),
            ("Total Price", "5000000 (Optional)", "emerald"),
            ("Status", "available (Optional - available, booked, sold, rented, under_construction, inactive)", "emerald"),
        ],
        "Houses": [
            ("Town Name", "Green Valley (Required)", "emerald"),
            ("Block Name", "Block A (Required)", "emerald"),
            ("Unit Number", "H-101 (Required)", "emerald"),
            ("Size", "10 Marla (Optional)", "emerald"),
            ("Total Price", "12000000 (Optional)", "emerald"),
            ("Status", "available (Optional - available, booked, sold, rented, under_construction, inactive)", "emerald"),
        ]
    }

    header_fills = {
        "blue": PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"),
        "emerald": PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid"),
        "purple": PatternFill(start_color="581C87", end_color="581C87", fill_type="solid"),
    }
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    hint_font = Font(name="Calibri", size=10, italic=True, color="71717A")

    for sheet_name, cols in sheets_config.items():
        ws = wb.create_sheet(title=sheet_name)

        # Headers
        headers = [c[0] for c in cols]
        hints = [c[1] for c in cols]
        color_key = cols[0][2]

        # Row 1: Headers
        ws.append(headers)
        # Row 2: Hints
        ws.append(hints)
        # Row 3: Sample data
        sample_values = []
        for c in cols:
            val = c[1].split(" (")[0]
            sample_values.append(val)
        ws.append(sample_values)

        # Format headers & hints
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fills[color_key]
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            hint_cell = ws.cell(row=2, column=col_idx)
            hint_cell.font = hint_font

        # Freeze top row
        ws.freeze_panes = "A2"

        # Set auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="REMS_Combined_Template.xlsx"'},
    )


@router.get("/templates/{module_key}")
def download_template(
    module_key: str,
    format: Literal["csv", "xlsx"] = "xlsx",
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager", "Staff", "Accountant")),
):
    handler = _get_handler(module_key)
    if not _user_can_import(user, handler.permission):
        raise HTTPException(403, "Insufficient permissions for this import module")
    content, media_type, filename = build_template(handler, format)
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/validate", response_model=ImportValidateResponse)
async def validate_import(
    module_key: str = Form(...),
    duplicate_mode: Literal["skip", "update", "create_only"] = Form("skip"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager", "Staff", "Accountant")),
):
    handler = _get_handler(module_key)
    if not _user_can_import(user, handler.permission):
        raise HTTPException(403, "Insufficient permissions")

    content = await file.read()
    rows, fmt = parse_upload(content, file.filename or "upload.csv")

    ctx = ImportContext(
        db=db,
        user_id=user.id,
        company_id=_company_id(user),
        duplicate_mode=duplicate_mode,
    )
    previews = validate_rows(module_key, rows, ctx)

    valid_count = sum(1 for p in previews if p["status"] == "valid")
    invalid_count = sum(1 for p in previews if p["status"] == "invalid")
    warning_count = sum(1 for p in previews if p["status"] == "warning")
    duplicate_count = sum(1 for p in previews if p["status"] == "duplicate")

    batch = create_batch(
        db, module_key, file.filename or "upload", fmt, duplicate_mode,
        len(rows), valid_count, user.id, _company_id(user),
    )
    _validation_cache[batch.id] = (rows, module_key, duplicate_mode)

    return ImportValidateResponse(
        module_key=module_key,
        file_name=file.filename or "upload",
        total_rows=len(rows),
        valid_count=valid_count,
        invalid_count=invalid_count,
        warning_count=warning_count,
        duplicate_count=duplicate_count,
        rows=[ImportRowPreview(**p) for p in previews],
        batch_id=batch.id,
    )


@router.post("/execute", response_model=ImportExecuteResponse)
def run_import(
    payload: ImportExecuteRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager")),
):
    handler = _get_handler(payload.module_key)
    if not _user_can_import(user, handler.permission):
        raise HTTPException(403, "Insufficient permissions")

    cached = _validation_cache.get(payload.batch_id) if payload.batch_id else None
    if not cached:
        raise HTTPException(400, "Validation session expired — please re-upload and validate")

    rows, module_key, duplicate_mode = cached
    if module_key != payload.module_key:
        raise HTTPException(400, "Module mismatch")

    batch = db.query(ImportBatch).filter(ImportBatch.id == payload.batch_id).first()
    if not batch:
        raise HTTPException(404, "Import batch not found")

    mode = payload.duplicate_mode or duplicate_mode
    ctx = ImportContext(db=db, user_id=user.id, company_id=_company_id(user), duplicate_mode=mode)

    row_numbers = payload.row_numbers
    if row_numbers is None:
        previews = validate_rows(module_key, rows, ctx)
        row_numbers = [p["row_number"] for p in previews if p["status"] == "valid"]

    counts = execute_import(db, module_key, rows, row_numbers, ctx, batch)

    batch.status = "completed" if counts["failed"] == 0 else "partial"
    batch.imported_rows = counts["imported"] + counts["updated"]
    batch.skipped_rows = counts["skipped"]
    batch.failed_rows = counts["failed"]
    from datetime import datetime
    batch.completed_at = datetime.utcnow()
    db.commit()

    if payload.batch_id in _validation_cache:
        del _validation_cache[payload.batch_id]

    return ImportExecuteResponse(
        batch_id=batch.id,
        status=batch.status,
        imported=counts["imported"],
        updated=counts["updated"],
        skipped=counts["skipped"],
        failed=counts["failed"],
        message=f"Import finished: {counts['imported']} created, {counts['updated']} updated.",
    )


@router.get("/history", response_model=list[ImportBatchOut])
def import_history(
    limit: int = 50,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager")),
):
    q = db.query(ImportBatch)
    cid = _company_id(user)
    if cid is not None:
        q = q.filter(ImportBatch.company_id == cid)
    q = q.order_by(ImportBatch.created_at.desc()).limit(limit)
    return q.all()


@router.get("/history/{batch_id}/errors")
def download_errors(
    batch_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager")),
):
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, "Batch not found")

    logs = (
        db.query(ImportRowLog)
        .filter(ImportRowLog.batch_id == batch_id, ImportRowLog.status.in_(["failed", "invalid"]))
        .order_by(ImportRowLog.row_number)
        .all()
    )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["row_number", "status", "error", "row_data"])
    for log in logs:
        writer.writerow([log.row_number, log.status, log.message or "", log.row_data_json or ""])

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="import_errors_{batch_id}.csv"'},
    )


def _execute_import_in_transaction(
    db: Session,
    module_key: str,
    raw_rows: list[dict[str, str]],
    row_numbers: list[int] | None,
    ctx: ImportContext,
    batch: ImportBatch,
) -> dict[str, int]:
    handler = _get_handler(module_key)
    valid_set = set(row_numbers) if row_numbers else None
    imported = updated = skipped = failed = 0

    for i, row in enumerate(raw_rows, start=2):
        if valid_set is not None and i not in valid_set:
            continue

        vr = handler.validate_row(row, ctx, i)
        if vr.status == "invalid":
            from app.services.bulk_import.engine import _log_row
            _log_row(db, batch.id, i, "failed", "; ".join(vr.errors), row)
            failed += 1
            continue

        nested = db.begin_nested()
        try:
            result = handler.import_row(row, ctx)
            if result.action == "created":
                imported += 1
                log_status = "imported"
            elif result.action == "updated":
                updated += 1
                log_status = "updated"
            elif result.action == "skipped":
                skipped += 1
                log_status = "skipped"
            else:
                failed += 1
                log_status = "failed"

            nested.commit()
            from app.services.bulk_import.engine import _log_row
            _log_row(db, batch.id, i, log_status, result.message, row, result.entity_type, result.entity_id)
        except Exception as exc:
            nested.rollback()
            failed += 1
            from app.services.bulk_import.engine import _log_row
            _log_row(db, batch.id, i, "failed", str(exc), row)

    if failed > 0:
        raise ValueError(f"Import failed for {failed} rows in module '{module_key}'. Rolling back entire batch.")

    return {"imported": imported, "updated": updated, "skipped": skipped, "failed": failed}


def _local_normalize_header(h: str) -> str:
    import re
    return re.sub(r"[^a-z0-9_]+", "_", (h or "").strip().lower()).strip("_")


def _local_cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "true" if val else "false"
    return str(val).strip()


def parse_xlsx_sheet_rows(content: bytes, sheet_name: str) -> list[dict[str, str]]:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return []

    target_sheet = None
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower():
            target_sheet = wb[name]
            break

    if target_sheet is None:
        wb.close()
        return []

    rows_iter = target_sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    headers = [_local_normalize_header(_local_cell_str(h)) for h in header_row]
    if not any(headers):
        wb.close()
        return []

    rows_to_process = []
    try:
        # Check if second row is hints
        second_row = next(rows_iter)
        is_hint = False
        for val in second_row:
            s_val = _local_cell_str(val)
            if "(Required)" in s_val or "(Optional)" in s_val:
                is_hint = True
                break

        if not is_hint:
            row = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                row[key] = _local_cell_str(second_row[idx] if idx < len(second_row) else None)
            if any(v for v in row.values()):
                rows_to_process.append(row)

        for cells in rows_iter:
            row = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                row[key] = _local_cell_str(cells[idx] if idx < len(cells) else None)

            # Skip template sample rows
            first_val = _local_cell_str(row.get(headers[0], ""))
            if first_val in ["John Doe", "Sunset Villa", "Ali Hassan"]:
                continue

            if any(v for v in row.values()):
                rows_to_process.append(row)
    except StopIteration:
        pass

    wb.close()
    return rows_to_process


def map_row_keys(row: dict[str, str], module_key: str) -> dict[str, str]:
    if module_key == "leads":
        return {
            "name": row.get("name", ""),
            "phone": row.get("phone", ""),
            "email": row.get("email", ""),
            "source": row.get("source", ""),
            "status": row.get("status", "") or "new",
            "assigned_to": row.get("assigned_to", ""),
        }
    elif module_key == "properties":
        return {
            "property_code": "",
            "name": row.get("property_title", "") or row.get("name", ""),
            "type": row.get("type", ""),
            "category": "Residential",
            "address": row.get("location", "") or row.get("address", ""),
            "size": "",
            "price": row.get("price", ""),
            "status": row.get("status", "") or "available",
            "for_sale": "true",
        }
    elif module_key == "employees":
        full_name = row.get("full_name", "") or row.get("name", "")
        parts = full_name.strip().split(" ", 1)
        first_name = parts[0] if parts else ""
        last_name = parts[1] if len(parts) > 1 else "-"
        return {
            "employee_id": "",
            "first_name": first_name,
            "last_name": last_name,
            "role": row.get("position", "") or row.get("role", ""),
            "department": row.get("department", ""),
            "phone": row.get("phone", ""),
            "work_email": "",
            "salary": row.get("salary", ""),
            "joining_date": datetime.utcnow().strftime("%Y-%m-%d"),
            "employment_status": "Active",
        }
    return row


@router.post("/master-validate", response_model=MasterValidateResponse)
async def master_validate(
    files: list[UploadFile] = File(...),
    duplicate_mode: Literal["skip", "update", "create_only"] = Form("skip"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager")),
):
    single_xlsx = False
    file_content = None

    if len(files) == 1 and files[0].filename.lower().endswith((".xlsx", ".xls")):
        file_content = await files[0].read()
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            sheet_names_lower = [s.lower() for s in wb.sheetnames]
            if any(s in sheet_names_lower for s in ["leads", "properties", "employees"]):
                single_xlsx = True
            wb.close()
        except Exception:
            pass

    results = {}
    ctx = ImportContext(
        db=db,
        user_id=user.id,
        company_id=_company_id(user),
        duplicate_mode=duplicate_mode,
    )

    if single_xlsx:
        for module_key in ["employees", "properties", "leads"]:
            raw_rows = parse_xlsx_sheet_rows(file_content, module_key)
            if not raw_rows:
                continue

            rows = [map_row_keys(r, module_key) for r in raw_rows]
            previews = validate_rows(module_key, rows, ctx)

            valid_count = sum(1 for p in previews if p["status"] == "valid")
            invalid_count = sum(1 for p in previews if p["status"] == "invalid")
            warning_count = sum(1 for p in previews if p["status"] == "warning")
            duplicate_count = sum(1 for p in previews if p["status"] == "duplicate")

            batch = create_batch(
                db, module_key, files[0].filename, "xlsx", duplicate_mode,
                len(rows), valid_count, user.id, _company_id(user),
            )
            _validation_cache[batch.id] = (rows, module_key, duplicate_mode)

            results[module_key] = ImportValidateResponse(
                module_key=module_key,
                file_name=files[0].filename,
                total_rows=len(rows),
                valid_count=valid_count,
                invalid_count=invalid_count,
                warning_count=warning_count,
                duplicate_count=duplicate_count,
                rows=[ImportRowPreview(**p) for p in previews],
                batch_id=batch.id,
            )
    else:
        extracted_files: list[tuple[str, bytes]] = []

        if len(files) == 1 and files[0].filename.lower().endswith(".zip"):
            zip_content = await files[0].read()
            try:
                with zipfile.ZipFile(io.BytesIO(zip_content)) as z:
                    for name in z.namelist():
                        if name.startswith("__") or "/" in name:
                            continue
                        if name.lower().endswith((".csv", ".xlsx")):
                            content = z.read(name)
                            extracted_files.append((name, content))
            except Exception as e:
                raise HTTPException(400, f"Invalid ZIP package: {str(e)}")
        else:
            for f in files:
                content = await f.read()
                extracted_files.append((f.filename, content))

        matched: dict[str, tuple[bytes, str]] = {}
        for filename, content in extracted_files:
            name_lower = filename.lower()
            if "employee" in name_lower:
                matched["employees"] = (content, filename)
            elif "property" in name_lower or "prop" in name_lower:
                matched["properties"] = (content, filename)
            elif "lead" in name_lower:
                matched["leads"] = (content, filename)

        if not matched:
            raise HTTPException(
                400,
                "No matching sheets or files found. Upload the REMS Combined Template workbook, or files/ZIP containing 'employee', 'property', or 'lead' in their names."
            )

        for module_key in ["employees", "properties", "leads"]:
            if module_key in matched:
                content, filename = matched[module_key]
                rows, fmt = parse_upload(content, filename)
                previews = validate_rows(module_key, rows, ctx)

                valid_count = sum(1 for p in previews if p["status"] == "valid")
                invalid_count = sum(1 for p in previews if p["status"] == "invalid")
                warning_count = sum(1 for p in previews if p["status"] == "warning")
                duplicate_count = sum(1 for p in previews if p["status"] == "duplicate")

                batch = create_batch(
                    db, module_key, filename, fmt, duplicate_mode,
                    len(rows), valid_count, user.id, _company_id(user),
                )
                _validation_cache[batch.id] = (rows, module_key, duplicate_mode)

                results[module_key] = ImportValidateResponse(
                    module_key=module_key,
                    file_name=filename,
                    total_rows=len(rows),
                    valid_count=valid_count,
                    invalid_count=invalid_count,
                    warning_count=warning_count,
                    duplicate_count=duplicate_count,
                    rows=[ImportRowPreview(**p) for p in previews],
                    batch_id=batch.id,
                )

    return MasterValidateResponse(
        employees=results.get("employees"),
        properties=results.get("properties"),
        leads=results.get("leads"),
        message=f"Validation completed. Loaded sheets/modules: {', '.join(results.keys())}"
    )


@router.post("/master-execute", response_model=MasterExecuteResponse)
def master_execute(
    payload: MasterExecuteRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager")),
):
    batches_to_run = []
    if payload.employees_batch_id:
        batches_to_run.append(("employees", payload.employees_batch_id))
    if payload.properties_batch_id:
        batches_to_run.append(("properties", payload.properties_batch_id))
    if payload.leads_batch_id:
        batches_to_run.append(("leads", payload.leads_batch_id))

    if not batches_to_run:
        raise HTTPException(400, "No batches specified for execution.")

    results = {}

    try:
        with db.begin_nested():
            for module_key, batch_id in batches_to_run:
                cached = _validation_cache.get(batch_id)
                if not cached:
                    raise HTTPException(400, f"Validation session expired for batch {batch_id} — please validate again")

                rows, cache_key, duplicate_mode = cached
                if cache_key != module_key:
                    raise HTTPException(400, "Module mismatch")

                batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
                if not batch:
                    raise HTTPException(404, f"Import batch {batch_id} not found")

                mode = payload.duplicate_mode or duplicate_mode
                ctx = ImportContext(db=db, user_id=user.id, company_id=_company_id(user), duplicate_mode=mode)

                previews = validate_rows(module_key, rows, ctx)
                row_numbers = [p["row_number"] for p in previews if p["status"] == "valid"]

                counts = _execute_import_in_transaction(db, module_key, rows, row_numbers, ctx, batch)

                batch.status = "completed"
                batch.imported_rows = counts["imported"] + counts["updated"]
                batch.skipped_rows = counts["skipped"]
                batch.failed_rows = 0
                batch.completed_at = datetime.utcnow()

                if batch_id in _validation_cache:
                    del _validation_cache[batch_id]

                results[module_key] = ImportExecuteResponse(
                    batch_id=batch.id,
                    status=batch.status,
                    imported=counts["imported"],
                    updated=counts["updated"],
                    skipped=counts["skipped"],
                    failed=0,
                    message=f"Import successful: {counts['imported']} created, {counts['updated']} updated."
                )

        db.commit()

    except Exception as e:
        db.rollback()
        for module_key, batch_id in batches_to_run:
            batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
            if batch:
                batch.status = "failed"
                batch.error_summary = str(e)
                db.add(batch)
        db.commit()
        raise HTTPException(400, f"Master execution rolled back: {str(e)}")

    return MasterExecuteResponse(
        employees=results.get("employees"),
        properties=results.get("properties"),
        leads=results.get("leads"),
        message="Master system update completed successfully."
    )


@router.post("/history/{batch_id}/rollback")
def rollback_import(
    batch_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    _: User = Depends(require_roles("Admin", "Manager")),
):
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, "Batch not found")

    if batch.status == "rolled_back":
        raise HTTPException(400, "This batch has already been rolled back.")

    logs = (
        db.query(ImportRowLog)
        .filter(ImportRowLog.batch_id == batch_id, ImportRowLog.status == "imported")
        .all()
    )

    from app.models.hr import Employee
    from app.models.property import Property
    from app.models.crm import Lead, Client, Dealer
    from app.models.finance import Account
    from app.models.tenant import Tenant
    from app.models.town import Town, Block, TownUnit

    model_map = {
        "employee": Employee,
        "property": Property,
        "lead": Lead,
        "client": Client,
        "crm_contact": Client,
        "dealer": Dealer,
        "account": Account,
        "tenant": Tenant,
        "town": Town,
        "block": Block,
        "town_unit": TownUnit,
    }

    deleted_count = 0
    errors = []

    try:
        with db.begin_nested():
            for log in logs:
                if not log.entity_type or not log.entity_id:
                    continue

                model_cls = model_map.get(log.entity_type.lower())
                if not model_cls:
                    continue

                record = db.query(model_cls).filter(model_cls.id == log.entity_id).first()
                if record:
                    try:
                        db.delete(record)
                        deleted_count += 1
                    except Exception as ex:
                        errors.append(f"Could not delete {log.entity_type} ID {log.entity_id}: {str(ex)}")

            if errors:
                raise ValueError("; ".join(errors))

        batch.status = "rolled_back"
        batch.error_summary = f"Rolled back by user {user.full_name} on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(400, f"Rollback failed due to references or constraints: {str(ex)}")

    return {"message": f"Successfully rolled back import. Deleted {deleted_count} records."}
