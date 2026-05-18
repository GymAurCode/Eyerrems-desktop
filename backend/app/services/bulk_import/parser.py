"""Parse CSV and XLSX upload files into row dicts."""
import csv
import io
import re
from typing import Any

from fastapi import HTTPException

MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 10_000
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def _normalize_header(h: str) -> str:
    return re.sub(r"[^a-z0-9_]+", "_", (h or "").strip().lower()).strip("_")


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "true" if val else "false"
    return str(val).strip()


def parse_upload(content: bytes, filename: str) -> tuple[list[dict[str, str]], str]:
    """Return (rows, format) where each row is column_key -> string value."""
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(400, f"File exceeds maximum size of {MAX_FILE_BYTES // (1024 * 1024)} MB")

    lower = filename.lower()
    ext = "." + lower.rsplit(".", 1)[-1] if "." in lower else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported file type. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}")

    if ext == ".csv":
        return _parse_csv(content), "csv"
    return _parse_xlsx(content), "xlsx"


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(400, "CSV has no header row")

    key_map = {_normalize_header(h): h for h in reader.fieldnames if h}
    rows: list[dict[str, str]] = []
    for i, raw in enumerate(reader):
        if i >= MAX_ROWS:
            raise HTTPException(400, f"Maximum {MAX_ROWS} rows per import")
        row: dict[str, str] = {}
        for norm, orig in key_map.items():
            row[norm] = _cell_str(raw.get(orig))
        if any(v for v in row.values()):
            rows.append(row)
    return rows


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(500, "XLSX support not available") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Workbook has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "Spreadsheet is empty")

    headers = [_normalize_header(_cell_str(h)) for h in header_row]
    if not any(headers):
        raise HTTPException(400, "Spreadsheet has no header row")

    out: list[dict[str, str]] = []
    for i, cells in enumerate(rows_iter):
        if i >= MAX_ROWS:
            raise HTTPException(400, f"Maximum {MAX_ROWS} rows per import")
        row = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            row[key] = _cell_str(cells[idx] if idx < len(cells) else None)
        if any(v for v in row.values()):
            out.append(row)
    wb.close()
    return out
