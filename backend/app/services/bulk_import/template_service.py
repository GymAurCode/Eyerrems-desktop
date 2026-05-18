"""Generate CSV/XLSX import templates."""
import csv
import io
from typing import Literal

from app.services.bulk_import.types import ColumnDef, ImportModuleHandler


def _header_row(columns: list[ColumnDef]) -> list[str]:
    return [c.label + ("*" if c.required else "") for c in columns]


def _sample_rows(columns: list[ColumnDef], count: int = 2) -> list[list[str]]:
    rows = []
    for i in range(count):
        rows.append([c.sample if i == 0 else "" for c in columns])
    return rows


def _notes_row(columns: list[ColumnDef]) -> list[str]:
    return [c.hint or "" for c in columns]


def build_csv_template(handler: ImportModuleHandler) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_header_row(handler.columns))
    writer.writerow([f"key:{c.key}" for c in handler.columns])
    writer.writerow(_notes_row(handler.columns))
    for row in _sample_rows(handler.columns):
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def build_xlsx_template(handler: ImportModuleHandler) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = handler.label[:31]

    headers = _header_row(handler.columns)
    ws.append(headers)
    ws.append([f"key:{c.key}" for c in handler.columns])
    ws.append(_notes_row(handler.columns))
    for row in _sample_rows(handler.columns, 2):
        ws.append(row)

    # Instructions sheet
    inst = wb.create_sheet("Instructions")
    inst.append(["Field Key", "Label", "Required", "Allowed Values", "Notes"])
    for c in handler.columns:
        inst.append([
            c.key,
            c.label,
            "Yes" if c.required else "No",
            ", ".join(c.enum_values) if c.enum_values else "",
            c.hint,
        ])

    header_font = Font(bold=True)
    fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_template(handler: ImportModuleHandler, fmt: Literal["csv", "xlsx"]) -> tuple[bytes, str, str]:
    if fmt == "csv":
        return build_csv_template(handler), "text/csv", f"{handler.key}_import_template.csv"
    return build_xlsx_template(handler), (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ), f"{handler.key}_import_template.xlsx"
