"""Report Engine — centralized generation of HTML, PDF, and Excel reports.

Usage:
    engine = ReportEngine(db)
    html = engine.generate("booking_statement", {"entity_id": 42}, "html")
    pdf_bytes = engine.generate("booking_statement", {"entity_id": 42}, "pdf")
    xlsx_bytes = engine.generate("booking_statement", {"entity_id": 42}, "xlsx")
"""
import base64
import io
import json
import logging
import mimetypes
import re
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.reports import ReportSettings

logger = logging.getLogger("rems.reports")

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = BASE_DIR / "templates" / "reports"

# ── Jinja2 environment ─────────────────────────────────────────────────────────
_env = Environment(
    loader=FileSystemLoader(str(TEMPLATE_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
)
_env.globals["max"] = max


# ── Registry: report_type → mapper function ────────────────────────────────────
_mappers: Dict[str, Callable] = {}


def register_report(report_type: str):
    """Decorator: register a data-mapper function for *report_type*."""
    def wrapper(fn: Callable):
        _mappers[report_type] = fn
        return fn
    return wrapper


def list_report_types() -> List[str]:
    return list(_mappers.keys())


def _strip_xhtml2pdf_unsupported(html: str) -> str:
    """Remove CSS features unsupported by xhtml2pdf's CSS parser.

    * @page rules containing margin boxes (like @bottom-left, @top-right etc.)
    * Entire @media blocks (contain nested braces)
    
    Keeps simple @page rules (size, margin) which xhtml2pdf handles fine.
    """

    def _find_block_end(text: str, start: int) -> int:
        """Find the closing brace index for a block starting at *start*,
        handling nested braces. Returns the index right after the closing ``}``.
        """
        j = start
        while j < len(text) and text[j] != "{":
            j += 1
        if j < len(text) and text[j] == "{":
            depth = 1
            j += 1
            while j < len(text):
                if text[j] == "{":
                    depth += 1
                elif text[j] == "}":
                    depth -= 1
                    if depth == 0:
                        j += 1
                        break
                j += 1
        return j

    def _block_contains_margin_boxes(text: str, block_start: int, block_end: int) -> bool:
        """Check if a @page block contains margin box selectors like @bottom-left."""
        markers = [
            "@top-left", "@top-center", "@top-right",
            "@bottom-left", "@bottom-center", "@bottom-right",
            "@left-top", "@left-middle", "@left-bottom",
            "@right-top", "@right-middle", "@right-bottom",
        ]
        block = text[block_start:block_end]
        for m in markers:
            if m in block:
                return True
        return False

    result = []
    i = 0
    while i < len(html):
        # Find the earliest occurrence of @page or @media
        page_pos = html.find("@page", i)
        media_pos = html.find("@media", i)

        if page_pos == -1 and media_pos == -1:
            result.append(html[i:])
            break

        if media_pos != -1 and (page_pos == -1 or media_pos < page_pos):
            # @media comes first — always strip the entire @media block
            result.append(html[i:media_pos])
            end = _find_block_end(html, media_pos + len("@media"))
            i = end
            continue

        if page_pos != -1:
            result.append(html[i:page_pos])
            block_end = _find_block_end(html, page_pos + len("@page"))
            if _block_contains_margin_boxes(html, page_pos, block_end):
                # Strip @page with margin boxes
                i = block_end
            else:
                # Keep simple @page (size, margin) — xhtml2pdf supports this
                result.append(html[page_pos:block_end])
                i = block_end
            continue

        result.append(html[i:])
        break

    return "".join(result)


# ── Core engine ────────────────────────────────────────────────────────────────

class ReportEngine:
    """Stateless report generator — instantiate per request with a DB session."""

    def __init__(self, db: Session):
        self.db = db

    # ── Public API ─────────────────────────────────────────────────────────────

    def generate(
        self,
        report_type: str,
        payload: dict,
        output_format: str = "html",
        options: Optional[dict] = None,
    ) -> Any:
        """Generate a report.
        Returns: HTML string, PDF bytes, or XLSX bytes depending on *output_format*.
        """
        options = options or {}

        # 1. Load settings
        settings = self._load_settings()

        # 1b. Merge per-report settings overrides (from workbench editing)
        settings_overrides = payload.get("settings_overrides") or {}
        if isinstance(settings_overrides, dict):
            for k, v in settings_overrides.items():
                if v is not None and hasattr(settings, k):
                    setattr(settings, k, v)

        # 2. Run mapper
        mapper = _mappers.get(report_type)
        if not mapper:
            raise ValueError(f"Unknown report type: {report_type}. Registered: {list_report_types()}")

        report_data = mapper(self.db, payload)
        if report_data is None:
            raise ValueError(f"Mapper for '{report_type}' returned no data")

        # 3. Render HTML
        template = _env.get_template("master_report.html")
        now = datetime.now(timezone.utc)
        show_seal = self._should_show_seal(settings, report_type)

        logo_data_uri = self._logo_data_uri(settings)
        has_logo = bool(logo_data_uri)
        show_watermark = has_logo and getattr(settings, "show_logo_watermark", True)

        # Report meta fields (letterhead, prepared for/by, note)
        import uuid
        report_ref = (
            f"REP-{report_type[:4].upper()}-"
            f"{now.strftime('%Y%m%d')}-"
            f"{uuid.uuid4().hex[:6].upper()}"
        )
        prepared_for = payload.get("prepared_for") or ""
        prepared_by = payload.get("prepared_by") or ""
        report_note = payload.get("note") or ""

        # Footer left text for print footer
        footer_left_parts = [settings.company_name or "", report_data.letterhead.title or ""]
        if report_data.letterhead.reference_no:
            footer_left_parts.append(report_data.letterhead.reference_no)
        footer_left = " | ".join(p for p in footer_left_parts if p)

        # Watermark opacity (configurable via settings_overrides or default)
        watermark_opacity = settings_overrides.get("watermark_opacity", "0.06")

        html = template.render(
            settings=settings,
            data=report_data,
            now=now,
            show_seal=show_seal,
            monogram_text=self._monogram(settings),
            logo_data_uri=logo_data_uri,
            has_logo=has_logo,
            show_watermark=show_watermark,
            watermark_opacity=watermark_opacity,
            paper_size=options.get("paper_size") or settings.default_paper_size,
            orientation=options.get("orientation") or settings.default_orientation,
            footer_left=footer_left,
            report_ref=report_ref,
            prepared_for=prepared_for,
            prepared_by=prepared_by,
            report_note=report_note,
        )

        if output_format == "html":
            return html

        if output_format == "pdf":
            return self._html_to_pdf(html, options)

        if output_format == "xlsx":
            return self._generate_xlsx(report_data, settings, options)

        raise ValueError(f"Unsupported output format: {output_format}")

    def _logo_data_uri(self, db_settings_row) -> str:
        """Convert the stored logo file into a base64 data URI for reliable
        rendering in HTML, PDF (WeasyPrint), and print contexts."""
        url = getattr(db_settings_row, "logo_url", "") or ""
        if not url:
            return ""

        try:
            base = Path(settings.upload_dir).resolve()
        except Exception:
            base = Path("uploads").resolve()

        # Handle both relative (/uploads/...) and absolute URLs
        if "/uploads/" in url:
            rel = url[url.index("/uploads/") + len("/uploads/"):]
            full_path = base / rel
        else:
            full_path = Path(url)

        if not full_path.exists():
            return ""

        try:
            raw = full_path.read_bytes()
            mime_type, _ = mimetypes.guess_type(str(full_path))
            if not mime_type:
                mime_type = "image/png"
            return f"data:{mime_type};base64,{base64.b64encode(raw).decode('ascii')}"
        except Exception:
            return ""

    def generate_filename(self, report_type: str, payload: dict) -> str:
        """Generate a standard filename: {ReportType}_{ReferenceNo}_{YYYYMMDD}.{ext}"""
        now = datetime.now(timezone.utc).strftime("%Y%m%d")
        mapper = _mappers.get(report_type)
        ref = ""
        if mapper:
            try:
                data = mapper(self.db, payload)
                if data and data.letterhead.reference_no:
                    ref = f"_{data.letterhead.reference_no}"
            except Exception:
                pass
        return f"{report_type}{ref}_{now}"

    # ── Internal helpers ───────────────────────────────────────────────────────

    def _load_settings(self) -> Any:
        """Load or create default ReportSettings for the current company."""
        company_id = getattr(self.db, "company_id", None) or 1
        row = self.db.query(ReportSettings).filter(
            ReportSettings.company_id == company_id
        ).first()
        if not row:
            row = ReportSettings(company_id=company_id, company_name="Company Name")
            self.db.add(row)
            self.db.commit()
            self.db.refresh(row)
        return row

    def _should_show_seal(self, settings, report_type: str) -> bool:
        config = settings.show_seal_config
        if not config:
            return True
        try:
            cfg = json.loads(config)
            return cfg.get(report_type, True)
        except (json.JSONDecodeError, TypeError):
            return True

    def _monogram(self, settings) -> str:
        name = settings.company_name or ""
        parts = name.strip().split()
        if len(parts) >= 2:
            return (parts[0][0] + parts[1][0]).upper()
        return name[:2].upper() if name else "RE"

    def _html_to_pdf(self, html: str, options: dict) -> bytes:
        """Convert HTML to PDF — tries WeasyPrint first, then xhtml2pdf."""
        from io import StringIO

        # 1. WeasyPrint
        try:
            from weasyprint import HTML as WeasyHTML
            return WeasyHTML(string=html).write_pdf()
        except Exception:
            logger.debug("WeasyPrint unavailable", exc_info=True)

        # 2. xhtml2pdf
        try:
            from xhtml2pdf import pisa
        except ImportError:
            raise RuntimeError(
                "No PDF library available — install weasyprint or xhtml2pdf"
            )

        stripped = _strip_xhtml2pdf_unsupported(html)

        buf = io.BytesIO()
        try:
            pdf = pisa.CreatePDF(StringIO(stripped), dest=buf)
        except Exception as exc:
            raise RuntimeError(
                f"xhtml2pdf conversion failed: {exc}"
            ) from exc

        if pdf.err:
            raise RuntimeError(f"xhtml2pdf conversion failed: {pdf.err}")
        buf.seek(0)
        return buf.read()

    def _generate_xlsx(self, report_data, settings, options) -> bytes:
        """Generate a formatted .xlsx workbook from report data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Styles
        header_font = Font(bold=True, size=10)
        title_font = Font(bold=True, size=12)
        total_font = Font(bold=True, size=10)
        normal_font = Font(size=9)
        thin_border = Border(
            bottom=Side(style="thin"),
            top=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font_white = Font(bold=True, size=9, color="FFFFFF")

        row = 1

        # ── Summary block ──────────────────────────────────────────────────
        ws.cell(row=row, column=1, value=report_data.letterhead.title).font = title_font
        row += 1
        if report_data.letterhead.reference_no:
            ws.cell(row=row, column=1, value=f"Ref: {report_data.letterhead.reference_no}").font = normal_font
            row += 1

        row += 1  # blank row

        # ── Financial strip ────────────────────────────────────────────────
        if report_data.financial_strip:
            strip_row = row
            for ci, cell in enumerate(report_data.financial_strip):
                c = ws.cell(row=strip_row, column=ci + 1, value=cell.label)
                c.font = Font(bold=True, size=9)
            row += 1
            for ci, cell in enumerate(report_data.financial_strip):
                c = ws.cell(row=row, column=ci + 1, value=cell.value)
                c.font = Font(bold=True, size=10)
                if cell.inverted:
                    c.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    c.font = Font(bold=True, size=10, color="FFFFFF")
            row += 2

        # ── Ledger table ───────────────────────────────────────────────────
        if report_data.ledger.columns and report_data.ledger.rows:
            # Header row
            for ci, col in enumerate(report_data.ledger.columns):
                c = ws.cell(row=row, column=ci + 1, value=col.label)
                c.font = header_font_white
                c.fill = header_fill
                c.border = thin_border
            row += 1

            # Data rows
            for lr in report_data.ledger.rows:
                for ci, cell_val in enumerate(lr.cells):
                    c = ws.cell(row=row, column=ci + 1, value=cell_val)
                    c.font = total_font if lr.is_total else (
                        Font(bold=True, size=9) if lr.is_milestone else normal_font
                    )
                    if lr.is_total:
                        c.border = Border(top=Side(style="medium"))
                row += 1

            # Totals row
            if report_data.ledger.totals_row:
                for ci, cell_val in enumerate(report_data.ledger.totals_row.cells):
                    c = ws.cell(row=row, column=ci + 1, value=cell_val)
                    c.font = total_font
                    c.border = Border(top=Side(style="medium"))
                row += 1

        # ── Auto-width columns ─────────────────────────────────────────────
        for col_idx in range(1, (len(report_data.ledger.columns) or 5) + 1):
            max_len = 10
            for r in range(1, row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # ── Freeze header row ──────────────────────────────────────────────
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
