"""
Excel Generator — Enterprise-grade Excel export for reports.

Uses openpyxl for professional Excel generation with:
  - Bold headers
  - Auto-width columns
  - Formatted cells (currency, date, number)
  - Summary section
  - Totals row
  - Professional styling
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.services.reports.report_engine import ReportColumn, ReportMeta, ReportResult, ReportSummary


# ── Color constants ───────────────────────────────────────────────────────────
HEADER_BG = "1E3A5F"       # Dark blue header
HEADER_FG = "FFFFFF"       # White text
SUBHEADER_BG = "2E5090"    # Medium blue
ALT_ROW_BG = "F5F7FA"      # Light gray alternate rows
TOTAL_BG = "E8F0FE"        # Light blue totals
BORDER_COLOR = "D0D7E3"    # Light border
SUMMARY_BG = "EBF5FB"      # Summary card background


def _thin_border():
    """Create thin border style."""
    thin = Side(style="thin", color=BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(ws, row: int, col: int, value: str, bold: bool = True):
    """Apply header styling to a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, color=HEADER_FG, size=10)
    cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()
    return cell


def _data_cell(ws, row: int, col: int, value: Any, align: str = "left", bold: bool = False, bg: Optional[str] = None):
    """Apply data cell styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    return cell


def generate_excel_report(result: ReportResult) -> bytes:
    """
    Generate Excel from ReportResult.
    Returns Excel bytes ready for download.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = result.meta.title[:31]  # Excel sheet name limit

    current_row = 1

    # ── Report Header ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:{get_column_letter(max(len(result.columns), 1))}{current_row}")
    title_cell = ws.cell(row=current_row, column=1, value=result.meta.company_name or "Real Estate Management System")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    ws.merge_cells(f"A{current_row}:{get_column_letter(max(len(result.columns), 1))}{current_row}")
    report_title_cell = ws.cell(row=current_row, column=1, value=result.meta.title)
    report_title_cell.font = Font(name="Calibri", bold=True, size=12, color=SUBHEADER_BG)
    report_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    if result.meta.subtitle:
        ws.merge_cells(f"A{current_row}:{get_column_letter(max(len(result.columns), 1))}{current_row}")
        subtitle_cell = ws.cell(row=current_row, column=1, value=result.meta.subtitle)
        subtitle_cell.font = Font(name="Calibri", italic=True, size=10, color="666666")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    # Generated info
    gen_text = f"Generated: {result.meta.generated_at.strftime('%Y-%m-%d %H:%M')} | By: {result.meta.generated_by}"
    if result.meta.report_id:
        gen_text += f" | Report ID: {result.meta.report_id}"
    ws.merge_cells(f"A{current_row}:{get_column_letter(max(len(result.columns), 1))}{current_row}")
    gen_cell = ws.cell(row=current_row, column=1, value=gen_text)
    gen_cell.font = Font(name="Calibri", size=8, color="888888")
    gen_cell.alignment = Alignment(horizontal="center")
    current_row += 1

    # Filters applied
    if result.meta.filters_applied:
        filter_text = " | ".join([f"{k}: {v}" for k, v in result.meta.filters_applied.items()])
        ws.merge_cells(f"A{current_row}:{get_column_letter(max(len(result.columns), 1))}{current_row}")
        filter_cell = ws.cell(row=current_row, column=1, value=f"Filters: {filter_text}")
        filter_cell.font = Font(name="Calibri", italic=True, size=8, color="888888")
        filter_cell.alignment = Alignment(horizontal="center")
        current_row += 1

    current_row += 1  # Spacer

    # ── Summary Cards ─────────────────────────────────────────────────────────
    if result.summary:
        ws.cell(row=current_row, column=1, value="Summary").font = Font(
            name="Calibri", bold=True, size=10, color=HEADER_BG
        )
        current_row += 1

        for idx, card in enumerate(result.summary):
            col = (idx % 4) + 1
            if idx > 0 and idx % 4 == 0:
                current_row += 2

            label_cell = ws.cell(row=current_row, column=col, value=card.label)
            label_cell.font = Font(name="Calibri", size=8, color="666666")
            label_cell.fill = PatternFill(start_color=SUMMARY_BG, end_color=SUMMARY_BG, fill_type="solid")
            label_cell.alignment = Alignment(horizontal="center")

            value_str = _format_summary_value(card, result.meta.currency_symbol)
            value_cell = ws.cell(row=current_row + 1, column=col, value=value_str)
            value_cell.font = Font(name="Calibri", bold=True, size=11, color=HEADER_BG)
            value_cell.fill = PatternFill(start_color=SUMMARY_BG, end_color=SUMMARY_BG, fill_type="solid")
            value_cell.alignment = Alignment(horizontal="center")

        current_row += 3  # Move past summary cards

    current_row += 1  # Spacer

    # ── Table Header ──────────────────────────────────────────────────────────
    visible_cols = [col for col in result.columns if col.visible]
    if not visible_cols:
        return _save_workbook(wb)

    ws.row_dimensions[current_row].height = 20
    for col_idx, col in enumerate(visible_cols, start=1):
        _header_style(ws, current_row, col_idx, col.label)

    header_row = current_row
    current_row += 1

    # ── Data Rows ─────────────────────────────────────────────────────────────
    currency_cols = {col.key for col in visible_cols if col.data_type == "currency"}
    number_cols = {col.key for col in visible_cols if col.data_type == "number"}
    date_cols = {col.key for col in visible_cols if col.data_type == "date"}

    # Currency symbol for this report
    _sym = result.meta.currency_symbol

    def _parse_numeric(value) -> float:
        """Parse a value that may be a raw float or a pre-formatted currency string."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        # Strip any currency symbol and commas from pre-formatted strings
        cleaned = str(value).replace(",", "").replace(_sym, "").replace("$", "").replace("₨", "").strip()
        return float(cleaned) if cleaned else 0.0

    # Track totals for currency/number columns
    totals: Dict[str, float] = {col.key: 0.0 for col in visible_cols if col.data_type in ("currency", "number")}

    for row_idx, row in enumerate(result.rows):
        bg = ALT_ROW_BG if row_idx % 2 == 1 else None
        ws.row_dimensions[current_row].height = 16

        for col_idx, col in enumerate(visible_cols, start=1):
            value = row.get(col.key)
            align = col.align if col.align else "left"

            if col.key in currency_cols:
                try:
                    num_val = _parse_numeric(value)
                    cell = _data_cell(ws, current_row, col_idx, f"{_sym} {num_val:,.2f}", align="right", bg=bg)
                    totals[col.key] = totals.get(col.key, 0.0) + num_val
                except (TypeError, ValueError):
                    _data_cell(ws, current_row, col_idx, str(value or ""), align=align, bg=bg)
            elif col.key in number_cols:
                try:
                    num_val = _parse_numeric(value)
                    cell = _data_cell(ws, current_row, col_idx, num_val, align="right", bg=bg)
                    cell.number_format = '#,##0.00'
                    totals[col.key] = totals.get(col.key, 0.0) + num_val
                except (TypeError, ValueError):
                    _data_cell(ws, current_row, col_idx, str(value or ""), align=align, bg=bg)
            elif col.key in date_cols:
                if isinstance(value, datetime):
                    _data_cell(ws, current_row, col_idx, value.strftime("%Y-%m-%d"), align="center", bg=bg)
                else:
                    _data_cell(ws, current_row, col_idx, str(value or ""), align="center", bg=bg)
            else:
                _data_cell(ws, current_row, col_idx, str(value or ""), align=align, bg=bg)

        current_row += 1

    # ── Totals Row ────────────────────────────────────────────────────────────
    if totals and result.rows:
        ws.row_dimensions[current_row].height = 18
        for col_idx, col in enumerate(visible_cols, start=1):
            if col_idx == 1:
                cell = _data_cell(ws, current_row, col_idx, "TOTAL", bold=True, bg=TOTAL_BG)
            elif col.key in totals:
                if col.data_type == "currency":
                    cell = _data_cell(ws, current_row, col_idx, f"{_sym} {totals[col.key]:,.2f}", align="right", bold=True, bg=TOTAL_BG)
                else:
                    cell = _data_cell(ws, current_row, col_idx, totals[col.key], align="right", bold=True, bg=TOTAL_BG)
                    cell.number_format = '#,##0.00'
            else:
                _data_cell(ws, current_row, col_idx, "", bg=TOTAL_BG)
        current_row += 1

    # ── Auto-width Columns ────────────────────────────────────────────────────
    for col_idx, col in enumerate(visible_cols, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(col.label)

        for row in result.rows:
            value = row.get(col.key, "")
            cell_len = len(str(value or ""))
            if cell_len > max_length:
                max_length = cell_len

        # Cap width between 10 and 40 characters
        adjusted_width = min(max(max_length + 2, 10), 40)
        ws.column_dimensions[col_letter].width = adjusted_width

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return _save_workbook(wb)


def _format_summary_value(card: ReportSummary, currency_symbol: str = "₨") -> str:
    """Format summary card value."""
    if card.format == "currency":
        try:
            return f"{currency_symbol} {float(card.value):,.2f}"
        except (TypeError, ValueError):
            return str(card.value)
    elif card.format == "percentage":
        try:
            return f"{float(card.value):.1f}%"
        except (TypeError, ValueError):
            return str(card.value)
    return str(card.value)


def _save_workbook(wb) -> bytes:
    """Save workbook to bytes."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
